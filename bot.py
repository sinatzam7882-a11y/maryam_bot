import os
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters, CallbackQueryHandler
from groq import Groq

# ==================== تنظیمات لاگینگ ====================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ==================== تنظیمات ====================
BOT_TOKEN = os.getenv("BOT_TOKEN")
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
ADMIN_ID = 8065571732  # 🔴 این رو با ایدی خودت عوض کن!
CHANNEL_ID = "@synapdse_os"  # آیدی کانال

client = Groq(api_key=GROQ_API_KEY)

# فایل‌های JSON
USERS_FILE = "users.json"
SURVEY_FILE = "survey.json"
EXCEL_FILE = "users_data.xlsx"

# ==================== اطلاعات پشتیبانی ====================
SUPPORT_INFO = {
    "phone": "۰۹۱۲۳۴۵۶۷۸۹",
    "email": "support@businessbot.com",
    "telegram": "@BusinessBotSupport",
    "hours": "۹ الی ۱۸",
    "response_time": "حداکثر ۲۴ ساعت"
}

# ==================== توابع JSON ====================
def read_json(file_path, default={}):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return default
    return default

def write_json(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ==================== توابع ذخیره ====================
def save_user_info(user_id, info):
    users = read_json(USERS_FILE, {})
    if str(user_id) not in users:
        users[str(user_id)] = {}
    users[str(user_id)].update(info)
    users[str(user_id)]["telegram_id"] = str(user_id)  # ذخیره آیدی تلگرام
    users[str(user_id)]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(USERS_FILE, users)
    logger.info(f"✅ اطلاعات کاربر {user_id} ذخیره شد")
    return True

def save_survey_answer(user_id, section, answer):
    surveys = read_json(SURVEY_FILE, {})
    if str(user_id) not in surveys:
        surveys[str(user_id)] = {}
    surveys[str(user_id)][section] = answer
    surveys[str(user_id)]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(SURVEY_FILE, surveys)

def get_user_info(user_id):
    users = read_json(USERS_FILE, {})
    return users.get(str(user_id), {})

def is_user_registered(user_id):
    users = read_json(USERS_FILE, {})
    user_data = users.get(str(user_id), {})
    return user_data.get("first_name") is not None and user_data.get("first_name") != ""

def is_business_registered(user_id):
    users = read_json(USERS_FILE, {})
    user_data = users.get(str(user_id), {})
    return user_data.get("business_name") is not None and user_data.get("business_name") != ""

def is_member_of_channel(user_id, context):
    """بررسی عضویت کاربر در کانال"""
    try:
        chat_member = context.bot.get_chat_member(chat_id=CHANNEL_ID, user_id=user_id)
        return chat_member.status in ['member', 'administrator', 'creator']
    except:
        return False

# ==================== خروجی اکسل ====================
def generate_excel_report():
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "اطلاعات کاربران"
    
    headers = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "تاریخ تولد", 
               "شماره تماس", "شهر", "نام کسب و کار", "آدرس", "راه معرفی", "تاریخ ثبت"]
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, (user_id, info) in enumerate(users.items(), 2):
        ws1.cell(row=row, column=1, value=row-1)
        ws1.cell(row=row, column=2, value=info.get('telegram_id', user_id))
        ws1.cell(row=row, column=3, value=info.get('first_name', ''))
        ws1.cell(row=row, column=4, value=info.get('last_name', ''))
        ws1.cell(row=row, column=5, value=info.get('birth_date', ''))
        ws1.cell(row=row, column=6, value=info.get('phone', ''))
        ws1.cell(row=row, column=7, value=info.get('city', ''))
        ws1.cell(row=row, column=8, value=info.get('business_name', ''))
        ws1.cell(row=row, column=9, value=info.get('address', ''))
        ws1.cell(row=row, column=10, value=info.get('referral_source', ''))
        ws1.cell(row=row, column=11, value=info.get('last_update', ''))
    
    for col in range(1, len(headers)+1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    ws2 = wb.create_sheet("پرسشنامه")
    
    survey_headers = ["ردیف", "آیدی تلگرام", "نام", "نام خانوادگی", "نام کسب و کار",
                     "درباره کسب و کار", "محصولات و مزیت", "زیرساخت مجازی", 
                     "تیم", "فروش ماهیانه", "چالش اصلی", "نیاز مشاوره"]
    
    for col, header in enumerate(survey_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, (user_id, answers) in enumerate(surveys.items(), 2):
        user_info = users.get(user_id, {})
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=user_info.get('telegram_id', user_id))
        ws2.cell(row=row, column=3, value=user_info.get('first_name', ''))
        ws2.cell(row=row, column=4, value=user_info.get('last_name', ''))
        ws2.cell(row=row, column=5, value=user_info.get('business_name', ''))
        ws2.cell(row=row, column=6, value=answers.get('about_business', ''))
        ws2.cell(row=row, column=7, value=answers.get('products', ''))
        ws2.cell(row=row, column=8, value=answers.get('infrastructure', ''))
        ws2.cell(row=row, column=9, value=answers.get('team', ''))
        ws2.cell(row=row, column=10, value=answers.get('sales', ''))
        ws2.cell(row=row, column=11, value=answers.get('problem', ''))
        ws2.cell(row=row, column=12, value=answers.get('consulting', ''))
    
    for col in range(1, len(survey_headers)+1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    ws3 = wb.create_sheet("خلاصه آمار")
    
    stats_data = [
        ["آمار کلی", ""],
        ["تعداد کل کاربران", len(users)],
        ["تعداد پرسشنامه‌های تکمیل شده", sum(1 for u in surveys if len(surveys[u]) > 2)],
        ["تاریخ تولید گزارش", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["آخرین ۵ کاربر", ""],
    ]
    
    for i, (user_id, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        stats_data.append([f"{i}. {name}", info.get('business_name', '')])
    
    for row, (key, value) in enumerate(stats_data, 1):
        ws3.cell(row=row, column=1, value=key)
        ws3.cell(row=row, column=2, value=value)
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 30
    
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# ==================== منوهای جدید ====================
# منوی اصلی با ۶ گزینه جدید
main_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🟢 بازار کار"), KeyboardButton("🔵 کسب‌وکار")],
    [KeyboardButton("🟣 مسئولیت اجتماعی"), KeyboardButton("🟠 مسیر رشد")],
    [KeyboardButton("🔴 لیدی لجستیک"), KeyboardButton("🌱 محصولات سیناپس")],
    [KeyboardButton("📖 راهنمای انتخاب مسیر"), KeyboardButton("🆔 اطلاعات شخصی")],
    [KeyboardButton("🏢 اطلاعات کسب و کار"), KeyboardButton("📊 پرسشنامه تخصصی")],
    [KeyboardButton("💬 مشاوره هوشمند"), KeyboardButton("📞 ارتباط با پشتیبانی")]
], resize_keyboard=True)

# ساب‌منوی بازار کار
market_menu = ReplyKeyboardMarkup([
    [KeyboardButton("👤 کارجو"), KeyboardButton("💼 فریلنسر")],
    [KeyboardButton("🏢 کارفرما")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

# ساب‌منوی کسب‌وکار
business_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🌟 برند شخصی"), KeyboardButton("🚀 برند محصولی")],
    [KeyboardButton("🏛️ برند سازمانی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

# ساب‌منوی مسئولیت اجتماعی
social_menu = ReplyKeyboardMarkup([
    [KeyboardButton("❤️ نیک‌اندیش داخل ایران"), KeyboardButton("🌍 نیک‌اندیش خارج ایران")],
    [KeyboardButton("🤝 پروژه اجتماعی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

# ساب‌منوی مسیر رشد
growth_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🧠 توسعه فردی"), KeyboardButton("🎯 توسعه شغلی")],
    [KeyboardButton("📈 توسعه اثر اجتماعی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

# ساب‌منوی لیدی لجستیک
logistics_menu = ReplyKeyboardMarkup([
    [KeyboardButton("💰 استعلام قیمت"), KeyboardButton("🌍 تأمین‌کننده خارجی")],
    [KeyboardButton("📦 حمل و اسناد"), KeyboardButton("📈 فروش و بازاریابی")],
    [KeyboardButton("🎓 آموزش واردات"), KeyboardButton("🧭 مشاوره تخصصی")],
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

back_menu = ReplyKeyboardMarkup([
    [KeyboardButton("🔙 بازگشت به منوی اصلی")]
], resize_keyboard=True)

def get_confirm_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تایید و ثبت", callback_data="confirm")],
        [InlineKeyboardButton("✏️ ویرایش اطلاعات", callback_data="edit")],
        [InlineKeyboardButton("❌ انصراف", callback_data="cancel")]
    ])

# ==================== سوالات ====================
personal_info_questions = [
    ("first_name", "👤 نام خود را وارد کنید:"),
    ("last_name", "👨‍👩‍👧 نام خانوادگی خود را وارد کنید:"),
    ("birth_date", "📅 تاریخ تولد (مثال: 1370/05/15):"),
    ("phone", "📞 شماره تماس موبایل:"),
    ("city", "🏙️ شهر خود را وارد کنید:"),
]

business_info_questions = [
    ("business_name", "🏢 نام کسب و کار خود را وارد کنید:"),
    ("address", "📍 آدرس شعب / دفتر مرکزی:"),
    ("referral_source", "📢 از چه طریقی با ما آشنا شدید؟"),
]

survey_questions = [
    ("about_business", "1️⃣ درباره کسب و کار خود بنویسید:\n(فعالیت، سابقه، اهداف)"),
    ("products", "2️⃣ چه محصولات یا خدماتی دارید؟\nمزیت شما نسبت به رقبا چیست؟"),
    ("infrastructure", "3️⃣ چه زیرساخت‌های مجازی دارید؟\n(سایت، اینستاگرام، اپلیکیشن، ...)"),
    ("team", "4️⃣ تیم شما شامل چه نیروهایی است؟\n(تعداد و تخصص)"),
    ("sales", "5️⃣ بالاترین فروش ماهیانه شما چقدر بوده؟"),
    ("problem", "6️⃣ مهمترین چالش یا مشکل فعلی شما چیست؟"),
    ("consulting", "7️⃣ در چه زمینه‌ای نیاز به مشاوره دارید؟")
]

# ==================== وضعیت کاربران ====================
user_states = {}

def get_user_state(user_id):
    return user_states.get(user_id, {"section": None, "step": 0, "temp": {}})

def set_user_state(user_id, section, step=0, temp=None):
    user_states[user_id] = {
        "section": section,
        "step": step,
        "temp": temp if temp else {}
    }

def clear_user_state(user_id):
    if user_id in user_states:
        del user_states[user_id]

# ==================== نوتیفیکیشن به ادمین ====================
async def notify_admin(context, user_id, info, section_type="personal"):
    try:
        if section_type == "personal":
            message = f"🆕 **کاربر جدید ثبت‌نام کرد!**\n\n"
            message += f"🆔 آیدی تلگرام: `{user_id}`\n"
            message += f"👤 نام: {info.get('first_name', '')} {info.get('last_name', '')}\n"
            message += f"🏙️ شهر: {info.get('city', '')}\n"
            message += f"📞 شماره: {info.get('phone', '')}"
        else:
            message = f"🏢 **اطلاعات کسب و کار جدید!**\n\n"
            message += f"🆔 آیدی تلگرام: `{user_id}`\n"
            message += f"👤 کاربر: {info.get('first_name', '')} {info.get('last_name', '')}\n"
            message += f"🏢 نام کسب و کار: {info.get('business_name', '')}\n"
            message += f"📍 آدرس: {info.get('address', '')}"
        
        await context.bot.send_message(chat_id=ADMIN_ID, text=message, parse_mode='Markdown')
    except Exception as e:
        logger.error(f"خطا در ارسال نوتیفیکیشن: {e}")

# ==================== دستور start ====================
async def start(update: Update, context):
    user_id = update.effective_user.id
    
    # بررسی عضویت در کانال
    if not is_member_of_channel(user_id, context):
        await update.message.reply_text(
            f"⚠️ **برای استفاده از ربات، ابتدا باید در کانال ما عضو شوید!**\n\n"
            f"📢 لطفاً ابتدا در کانال [{CHANNEL_ID}](https://t.me/{CHANNEL_ID[1:]}) عضو شوید.\n\n"
            f"✅ سپس دوباره روی /start کلیک کنید.",
            parse_mode='Markdown',
            disable_web_page_preview=True
        )
        return
    
    user_info = get_user_info(user_id)
    logger.info(f"کاربر {user_id} وارد شد")
    
    if is_user_registered(user_id):
        welcome_msg = f"""✨ خوش برگشتی {user_info.get('first_name')} عزیز!

به سیناپس خوش اومدی. 🌱😍
هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. تو از کجا میخوای شروع کنی؟

🟢 بازار کار
🔵 کسب‌وکار
🟣 مسئولیت اجتماعی
🟠 مسیر رشد
🔴 لیدی لجستیک
🌱 محصولات سیناپس

لطفاً مسیر موردنظرت را انتخاب کن. 👇"""
    else:
        welcome_msg = """سلام سلام
شهبازی هستم، مریم 😍🌱
اینجا قراره هویت کسب و کار و برند خودتون رو بسازید و روز به روز فروش بیشتری رو تجربه کنین. 
با من همراه باش

به سیناپس خوش اومدی. 🌱😍
هر آدمی در یکی از این مسیرها به دنبال رشد و توسعه برای ساختن یک ورژن بهتر از خودشه. تو از کجا میخوای شروع کنی؟

🟢 بازار کار
🔵 کسب‌وکار
🟣 مسئولیت اجتماعی
🟠 مسیر رشد
🔴 لیدی لجستیک
🌱 محصولات سیناپس

لطفاً مسیر موردنظرت را انتخاب کن. 👇"""

    try:
        with open('images/welcome.jpg', 'rb') as photo:
            await update.message.reply_photo(
                photo=photo,
                caption=welcome_msg,
                reply_markup=main_menu
            )
    except:
        await update.message.reply_text(welcome_msg, reply_markup=main_menu)

# ==================== تابع نمایش خلاصه ====================
def get_info_summary(info, section_type):
    if section_type == "personal":
        return f"""📝 **خلاصه اطلاعات شخصی شما:**

👤 نام: {info.get('first_name', '❌')}
👨‍👩‍👧 نام خانوادگی: {info.get('last_name', '❌')}
📅 تاریخ تولد: {info.get('birth_date', '❌')}
📞 شماره تماس: {info.get('phone', '❌')}
🏙️ شهر: {info.get('city', '❌')}

آیا اطلاعات صحیح است؟"""
    
    elif section_type == "business":
        return f"""🏢 **خلاصه اطلاعات کسب و کار شما:**

نام کسب و کار: {info.get('business_name', '❌')}
📍 آدرس: {info.get('address', '❌')}
📢 راه معرفی: {info.get('referral_source', '❌')}

آیا اطلاعات صحیح است؟"""
    
    return ""

# ==================== پردازش منو ====================
async def handle_menu(update: Update, context):
    user_id = update.effective_user.id
    text = update.message.text
    
    # بررسی عضویت در کانال
    if not is_member_of_channel(user_id, context):
        await update.message.reply_text(
            f"⚠️ **برای استفاده از ربات، ابتدا باید در کانال ما عضو شوید!**\n\n"
            f"📢 لطفاً ابتدا در کانال [{CHANNEL_ID}](https://t.me/{CHANNEL_ID[1:]}) عضو شوید.\n\n"
            f"✅ سپس دوباره روی /start کلیک کنید.",
            parse_mode='Markdown',
            disable_web_page_preview=True
        )
        return
    
    # ========== راهنمای انتخاب مسیر ==========
    if text == "📖 راهنمای انتخاب مسیر":
        guide_text = """📖 **راهنمای انتخاب مسیر**

به سیناپس خوش اومدی. 🌱

🟢 **بازار کار**
اگر به دنبال شغل، پروژه یا درآمد یا جذب نیرو هستی.

🔵 **کسب‌وکار**
اگر می‌خواهی کسب‌وکارت را رشد بدهی یا مسئله‌ای را حل کنی.

🟣 **مسئولیت اجتماعی**
اگر می‌خواهی در کنار رشد خودت، اثری مثبت بر جامعه بگذاری.

🟠 **مسیر رشد**
اگر به دنبال خودشناسی و طراحی مسیر زندگی و شغلی هستی.

🔴 **لیدی لجستیک**
اگر می‌خواهی با فرصت‌ها و خدمات اکوسیستم لیدی لجستیک همراه شوی.

🌱 **محصولات سیناپس**
اگر می‌خواهی از ابزارها، دوره‌ها و خدمات سیناپس استفاده کنی.

مسیر موردنظرت را انتخاب کن. 👇"""
        
        await update.message.reply_text(guide_text, reply_markup=main_menu, parse_mode='Markdown')
        return
    
    # ========== بازار کار ==========
    if text == "🟢 بازار کار":
        await update.message.reply_text(
            "🟢 **بازار کار**\n\n"
            "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=market_menu
        )
        return
    
    if text in ["👤 کارجو", "💼 فریلنسر", "🏢 کارفرما"]:
        await update.message.reply_text(
            f"✅ **شما گزینه '{text}' را انتخاب کردید.**\n\n"
            f"به زودی خدمات مربوط به این بخش در دسترس قرار می‌گیرد.\n"
            f"برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=market_menu
        )
        return
    
    # ========== کسب‌وکار ==========
    if text == "🔵 کسب‌وکار":
        await update.message.reply_text(
            "🔵 **کسب‌وکار**\n\n"
            "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=business_menu
        )
        return
    
    if text in ["🌟 برند شخصی", "🚀 برند محصولی", "🏛️ برند سازمانی"]:
        await update.message.reply_text(
            f"✅ **شما گزینه '{text}' را انتخاب کردید.**\n\n"
            f"به زودی خدمات مربوط به این بخش در دسترس قرار می‌گیرد.\n"
            f"برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=business_menu
        )
        return
    
    # ========== مسئولیت اجتماعی ==========
    if text == "🟣 مسئولیت اجتماعی":
        await update.message.reply_text(
            "🟣 **مسئولیت اجتماعی**\n\n"
            "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=social_menu
        )
        return
    
    if text in ["❤️ نیک‌اندیش داخل ایران", "🌍 نیک‌اندیش خارج ایران", "🤝 پروژه اجتماعی"]:
        await update.message.reply_text(
            f"✅ **شما گزینه '{text}' را انتخاب کردید.**\n\n"
            f"به زودی خدمات مربوط به این بخش در دسترس قرار می‌گیرد.\n"
            f"برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=social_menu
        )
        return
    
    # ========== مسیر رشد ==========
    if text == "🟠 مسیر رشد":
        await update.message.reply_text(
            "🟠 **مسیر رشد**\n\n"
            "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=growth_menu
        )
        return
    
    if text in ["🧠 توسعه فردی", "🎯 توسعه شغلی", "📈 توسعه اثر اجتماعی"]:
        await update.message.reply_text(
            f"✅ **شما گزینه '{text}' را انتخاب کردید.**\n\n"
            f"به زودی خدمات مربوط به این بخش در دسترس قرار می‌گیرد.\n"
            f"برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=growth_menu
        )
        return
    
    # ========== لیدی لجستیک ==========
    if text == "🔴 لیدی لجستیک":
        await update.message.reply_text(
            "🔴 **لیدی لجستیک**\n\n"
            "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=logistics_menu
        )
        return
    
    if text in ["💰 استعلام قیمت", "🌍 تأمین‌کننده خارجی", "📦 حمل و اسناد", 
                "📈 فروش و بازاریابی", "🎓 آموزش واردات", "🧭 مشاوره تخصصی"]:
        await update.message.reply_text(
            f"✅ **شما گزینه '{text}' را انتخاب کردید.**\n\n"
            f"به زودی خدمات مربوط به این بخش در دسترس قرار می‌گیرد.\n"
            f"برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=logistics_menu
        )
        return
    
    # ========== محصولات سیناپس ==========
    if text == "🌱 محصولات سیناپس":
        await update.message.reply_text(
            "🌱 **محصولات سیناپس**\n\n"
            "به زودی ابزارها، دوره‌ها و خدمات سیناپس در این بخش قرار می‌گیرند.\n"
            "برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید.",
            reply_markup=main_menu
        )
        return
    
    # ========== بازگشت به منوی اصلی ==========
    if text == "🔙 بازگشت به منوی اصلی":
        clear_user_state(user_id)
        await update.message.reply_text(
            "🔹 به منوی اصلی بازگشتید 👇",
            reply_markup=main_menu
        )
        return
    
    # ========== اطلاعات شخصی ==========
    if text == "🆔 اطلاعات شخصی":
        if is_user_registered(user_id):
            user_info = get_user_info(user_id)
            await update.message.reply_text(
                f"✅ **شما قبلاً ثبت‌نام کرده‌اید!**\n\n"
                f"🆔 آیدی تلگرام: `{user_id}`\n"
                f"👤 نام: {user_info.get('first_name', '')} {user_info.get('last_name', '')}\n"
                f"🏙️ شهر: {user_info.get('city', '')}\n"
                f"📞 شماره: {user_info.get('phone', '')}\n\n"
                f"برای ویرایش اطلاعات، اطلاعات جدید را وارد کنید 👇",
                reply_markup=back_menu
            )
            clear_user_state(user_id)
            set_user_state(user_id, "personal", 0, {})
            await update.message.reply_text(
                f"✏️ **ویرایش اطلاعات شخصی**\n\n{personal_info_questions[0][1]}",
                parse_mode='Markdown'
            )
            return
        
        clear_user_state(user_id)
        set_user_state(user_id, "personal", 0, {})
        await update.message.reply_text(
            f"📝 **ثبت اطلاعات شخصی**\n\n{personal_info_questions[0][1]}",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
        return
    
    # ========== اطلاعات کسب و کار ==========
    if text == "🏢 اطلاعات کسب و کار":
        if not is_user_registered(user_id):
            await update.message.reply_text(
                "⚠️ لطفاً ابتدا در بخش '🆔 اطلاعات شخصی' ثبت‌نام کنید.",
                reply_markup=main_menu
            )
            return
        
        user_info = get_user_info(user_id)
        if is_business_registered(user_id):
            await update.message.reply_text(
                f"✅ **اطلاعات کسب و کار شما قبلاً ثبت شده:**\n\n"
                f"🏢 نام کسب و کار: {user_info.get('business_name', '')}\n"
                f"📍 آدرس: {user_info.get('address', '')}\n"
                f"📢 راه معرفی: {user_info.get('referral_source', '')}\n\n"
                f"برای ویرایش، اطلاعات جدید را وارد کنید 👇",
                reply_markup=back_menu
            )
            clear_user_state(user_id)
            set_user_state(user_id, "business", 0, {})
            await update.message.reply_text(
                f"✏️ **ویرایش اطلاعات کسب و کار**\n\n{business_info_questions[0][1]}",
                parse_mode='Markdown'
            )
            return
        
        clear_user_state(user_id)
        set_user_state(user_id, "business", 0, {})
        await update.message.reply_text(
            f"🏢 **ثبت اطلاعات کسب و کار**\n\n{business_info_questions[0][1]}",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
        return
    
    # ========== پرسشنامه ==========
    if text == "📊 پرسشنامه تخصصی":
        if not is_user_registered(user_id):
            await update.message.reply_text(
                "⚠️ لطفاً ابتدا در بخش '🆔 اطلاعات شخصی' ثبت‌نام کنید.",
                reply_markup=main_menu
            )
            return
        
        if not is_business_registered(user_id):
            await update.message.reply_text(
                "⚠️ لطفاً ابتدا اطلاعات کسب و کار خود را ثبت کنید.\n"
                "بخش '🏢 اطلاعات کسب و کار'",
                reply_markup=main_menu
            )
            return
        
        surveys = read_json(SURVEY_FILE, {})
        if str(user_id) in surveys and len(surveys[str(user_id)]) > 2:
            await update.message.reply_text(
                "✅ شما قبلاً پرسشنامه را تکمیل کرده‌اید.\n"
                "از بخش '💬 مشاوره هوشمند' می‌توانید سوالات خود را بپرسید.",
                reply_markup=main_menu
            )
            return
        
        clear_user_state(user_id)
        set_user_state(user_id, "survey", 0, {})
        await update.message.reply_text(
            f"📋 **پرسشنامه تخصصی**\n\n{survey_questions[0][1]}",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
        return
    
    # ========== مشاوره ==========
    if text == "💬 مشاوره هوشمند":
        if not is_user_registered(user_id):
            await update.message.reply_text(
                "⚠️ لطفاً ابتدا در بخش '🆔 اطلاعات شخصی' ثبت‌نام کنید.",
                reply_markup=main_menu
            )
            return
        
        await update.message.reply_text(
            "💬 **مشاوره هوشمند**\n\n"
            "سلام! من مریم هستم، مشاور کسب و کار شما.\n"
            "هر سوالی درباره برندسازی، بازاریابی، فروش و... داری بپرس.\n\n"
            "📌 برای بازگشت به منو، روی دکمه زیر کلیک کن 👇",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
        return
    
    # ========== پشتیبانی ==========
    if text == "📞 ارتباط با پشتیبانی":
        support_text = f"""📞 **ارتباط با پشتیبانی**

به تیم پشتیبانی ما خوش آمدید! 🎯

📱 **شماره تماس:** {SUPPORT_INFO['phone']}
📧 **ایمیل:** {SUPPORT_INFO['email']}
🆔 **آیدی تلگرام:** {SUPPORT_INFO['telegram']}
⏰ **ساعات پاسخگویی:** {SUPPORT_INFO['hours']}
⏱ **زمان پاسخگویی:** {SUPPORT_INFO['response_time']}

───────────────────
💬 **نکات مهم:**
• لطفاً قبل از تماس، شماره و نام خود را آماده داشته باشید
• در صورت نیاز به مشاوره فوری، از بخش 'مشاوره هوشمند' استفاده کنید
• پیام‌های پشتیبانی حداکثر ۲۴ ساعت پاسخ داده می‌شوند

برای بازگشت به منوی اصلی، روی دکمه زیر کلیک کنید 👇"""
        
        await update.message.reply_text(
            support_text,
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
        return
    
    # ========== پردازش پاسخ‌ها ==========
    state = get_user_state(user_id)
    section = state["section"]
    step = state["step"]
    temp = state["temp"]
    
    # پردازش اطلاعات شخصی
    if section == "personal":
        if step < len(personal_info_questions):
            field_name, _ = personal_info_questions[step]
            temp[field_name] = text
            
            if step + 1 < len(personal_info_questions):
                set_user_state(user_id, "personal", step + 1, temp)
                _, next_q = personal_info_questions[step + 1]
                await update.message.reply_text(next_q, reply_markup=back_menu)
            else:
                summary = get_info_summary(temp, "personal")
                set_user_state(user_id, "personal_confirm", 0, temp)
                await update.message.reply_text(
                    summary,
                    reply_markup=get_confirm_keyboard(),
                    parse_mode='Markdown'
                )
        return
    
    # پردازش اطلاعات کسب و کار
    if section == "business":
        if step < len(business_info_questions):
            field_name, _ = business_info_questions[step]
            temp[field_name] = text
            
            if step + 1 < len(business_info_questions):
                set_user_state(user_id, "business", step + 1, temp)
                _, next_q = business_info_questions[step + 1]
                await update.message.reply_text(next_q, reply_markup=back_menu)
            else:
                summary = get_info_summary(temp, "business")
                set_user_state(user_id, "business_confirm", 0, temp)
                await update.message.reply_text(
                    summary,
                    reply_markup=get_confirm_keyboard(),
                    parse_mode='Markdown'
                )
        return
    
    # پردازش پرسشنامه
    if section == "survey":
        if step < len(survey_questions):
            field_name, _ = survey_questions[step]
            temp[field_name] = text
            
            if step + 1 < len(survey_questions):
                set_user_state(user_id, "survey", step + 1, temp)
                _, next_q = survey_questions[step + 1]
                await update.message.reply_text(next_q, reply_markup=back_menu)
            else:
                for key, value in temp.items():
                    save_survey_answer(user_id, key, value)
                clear_user_state(user_id)
                
                await update.message.reply_text(
                    "🌹 **با تشکر از شما!** 🌹\n\n"
                    "پرسشنامه شما با موفقیت ثبت شد.\n"
                    "✅ **ظرف ۴۸ ساعت آینده کارشناسان ما با شما تماس می‌گیرند.**\n\n"
                    "🔹 به منوی اصلی بازگشتید 👇",
                    reply_markup=main_menu,
                    parse_mode='Markdown'
                )
        return
    
    # ========== گفتگو با Groq ==========
    if state["section"] is None:
        if not is_user_registered(user_id):
            await update.message.reply_text(
                "⚠️ لطفاً ابتدا در بخش '🆔 اطلاعات شخصی' ثبت‌نام کنید.",
                reply_markup=main_menu
            )
            return
        
        user_info = get_user_info(user_id)
        try:
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": f"""تو یک مشاور کسب و کار حرفه‌ای هستی به نام مریم شهبازی.
                    اطلاعات کاربر: نام: {user_info.get('first_name', '')} {user_info.get('last_name', '')}
                    کسب و کار: {user_info.get('business_name', '')}
                    با لحنی گرم و دوستانه پاسخ بده. حتما به فارسی روان پاسخ بده."""},
                    {"role": "user", "content": text}
                ],
                temperature=0.7,
                max_tokens=1000
            )
            await update.message.reply_text(
                response.choices[0].message.content,
                reply_markup=back_menu
            )
        except Exception as e:
            logger.error(f"خطا در Groq: {e}")
            await update.message.reply_text(
                "⚠️ خطا در ارتباط با سرور. لطفاً چند لحظه دیگر تلاش کنید.",
                reply_markup=back_menu
            )

# ==================== دکمه‌های اینلاین ====================
async def handle_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    data = query.data
    state = get_user_state(user_id)
    temp = state.get("temp", {})
    section = state.get("section", "")
    
    logger.info(f"کاربر {user_id} روی دکمه {data} کلیک کرد - بخش: {section}")
    
    if data == "confirm":
        try:
            if section == "personal_confirm":
                save_user_info(user_id, temp)
                await notify_admin(context, user_id, temp, "personal")
                
                await query.edit_message_reply_markup(reply_markup=None)
                
                await query.message.reply_text(
                    f"✅ **ثبت‌نام با موفقیت انجام شد!** 🎉\n\n"
                    f"🆔 آیدی تلگرام: `{user_id}`\n"
                    f"👤 نام: {temp.get('first_name', '')} {temp.get('last_name', '')}\n"
                    f"🏙️ شهر: {temp.get('city', '')}\n"
                    f"📞 شماره تماس: {temp.get('phone', '')}\n\n"
                    f"📌 اطلاعات شما در سیستم ثبت شد.\n"
                    f"به سیناپس خوش اومدی! 🌱😍\n\n"
                    f"از منوی زیر مسیر موردنظرت را انتخاب کن 👇",
                    reply_markup=main_menu,
                    parse_mode='Markdown'
                )
                clear_user_state(user_id)
                return
            
            elif section == "business_confirm":
                user_info = get_user_info(user_id)
                temp.update(user_info)
                save_user_info(user_id, temp)
                await notify_admin(context, user_id, temp, "business")
                
                await query.edit_message_reply_markup(reply_markup=None)
                
                await query.message.reply_text(
                    f"✅ **اطلاعات کسب و کار شما ثبت شد!** 🏢\n\n"
                    f"🏢 نام کسب و کار: {temp.get('business_name', '')}\n"
                    f"📍 آدرس: {temp.get('address', '')}\n"
                    f"📢 راه معرفی: {temp.get('referral_source', '')}\n\n"
                    f"📌 حالا می‌توانید:\n"
                    f"• پرسشنامه تخصصی را پر کنید\n"
                    f"• از مشاوره هوشمند استفاده کنید\n\n"
                    f"به منوی اصلی بازگشتید 👇",
                    reply_markup=main_menu,
                    parse_mode='Markdown'
                )
                clear_user_state(user_id)
                return
            
            else:
                await query.edit_message_reply_markup(reply_markup=None)
                await query.message.reply_text(
                    "⚠️ خطا در ثبت اطلاعات. لطفاً دوباره تلاش کنید.",
                    reply_markup=main_menu
                )
                clear_user_state(user_id)
                return
            
        except Exception as e:
            logger.error(f"خطا در تایید: {e}")
            await query.edit_message_reply_markup(reply_markup=None)
            await query.message.reply_text(
                f"⚠️ خطا در ثبت اطلاعات: {str(e)}\n\nلطفاً دوباره تلاش کنید.",
                reply_markup=main_menu
            )
            clear_user_state(user_id)
    
    elif data == "edit":
        if section == "personal_confirm":
            new_section = "personal"
            questions = personal_info_questions
            title = "✏️ ویرایش اطلاعات شخصی"
        elif section == "business_confirm":
            new_section = "business"
            questions = business_info_questions
            title = "✏️ ویرایش اطلاعات کسب و کار"
        else:
            await query.edit_message_text(
                "⚠️ خطا در ویرایش. لطفاً دوباره تلاش کنید.",
                reply_markup=main_menu
            )
            clear_user_state(user_id)
            return
        
        await query.edit_message_reply_markup(reply_markup=None)
        set_user_state(user_id, new_section, 0, {})
        await query.message.reply_text(
            f"{title}\n\n{questions[0][1]}",
            reply_markup=back_menu,
            parse_mode='Markdown'
        )
    
    elif data == "cancel":
        clear_user_state(user_id)
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text(
            "❌ **ثبت‌نامه لغو شد.**\n\n"
            "اطلاعات شما ذخیره نشد.\n"
            "در صورت تمایل می‌توانید دوباره ثبت‌نام کنید.\n\n"
            "به منوی اصلی بازگشتید 👇",
            reply_markup=main_menu,
            parse_mode='Markdown'
        )

# ==================== دستورات ادمین ====================
async def get_excel(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    await update.message.reply_text("📊 در حال تولید فایل اکسل... لطفاً صبر کنید...")
    
    try:
        excel_file = generate_excel_report()
        with open(excel_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f'users_data_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                caption="📊 **گزارش کامل کاربران**\n\n"
                       f"📅 تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                       "📋 شامل: اطلاعات کاربران + پرسشنامه + آمار"
            )
    except Exception as e:
        await update.message.reply_text(f"⚠️ خطا در تولید فایل: {str(e)}")

async def get_data(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'rb') as f:
            await update.message.reply_document(f, filename=f'users_{datetime.now().strftime("%Y%m%d")}.json')
    
    if os.path.exists(SURVEY_FILE):
        with open(SURVEY_FILE, 'rb') as f:
            await update.message.reply_document(f, filename=f'survey_{datetime.now().strftime("%Y%m%d")}.json')

async def show_summary(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    users = read_json(USERS_FILE, {})
    surveys = read_json(SURVEY_FILE, {})
    
    completed_surveys = sum(1 for u in surveys if len(surveys[u]) > 2)
    today = datetime.now().strftime("%Y-%m-%d")
    today_users = sum(1 for u in users.values() if u.get("last_update", "").startswith(today))
    
    summary = f"📊 **آمار کلی:**\n\n"
    summary += f"👥 کل کاربران: {len(users)}\n"
    summary += f"📋 پرسشنامه‌های تکمیل شده: {completed_surveys}\n"
    summary += f"🆕 کاربران امروز: {today_users}\n\n"
    summary += "**آخرین ۵ کاربر:**\n"
    
    for i, (uid, info) in enumerate(list(users.items())[-5:], 1):
        name = f"{info.get('first_name', '')} {info.get('last_name', '')}"
        business = info.get('business_name', 'نامشخص')
        summary += f"{i}. {name} - {business}\n"
    
    await update.message.reply_text(summary, parse_mode='Markdown')

async def broadcast(update: Update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ دسترسی ندارید!")
        return
    
    users = read_json(USERS_FILE, {})
    if not users:
        await update.message.reply_text("📭 هیچ کاربری ثبت نشده!")
        return
    
    message_text = " ".join(context.args)
    if not message_text:
        await update.message.reply_text("⚠️ لطفاً پیام خود را وارد کنید.\nمثال: /broadcast سلام به همه!")
        return
    
    await update.message.reply_text(f"📤 ارسال پیام به {len(users)} کاربر...")
    
    success = 0
    for user_id in users.keys():
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"📢 **پیام از طرف مدیریت:**\n\n{message_text}"
            )
            success += 1
        except:
            pass
    
    await update.message.reply_text(f"✅ پیام به {success} کاربر ارسال شد.")

# ==================== اجرا ====================
app = ApplicationBuilder().token(BOT_TOKEN).build()

app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("getexcel", get_excel))
app.add_handler(CommandHandler("getdata", get_data))
app.add_handler(CommandHandler("summary", show_summary))
app.add_handler(CommandHandler("broadcast", broadcast))
app.add_handler(CallbackQueryHandler(handle_callback))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu))

print("🤖 بات سیناپس با قابلیت‌های جدید روشن شد...")
print(f"📢 کانال اجباری: {CHANNEL_ID}")
print("📁 اطلاعات در فایل‌های JSON ذخیره می‌شوند")
print(f"👑 ادمین: {ADMIN_ID}")
print("📊 دستور /getexcel برای دریافت فایل اکسل")
app.run_polling()